import openpyxl
from openpyxl import Workbook

# Define events
events = [
    "Make", "Odyssey", "Design", "Pitching", "Quiz", "Gaming",
    "GD", "Programming", "Surprise", "CryptX", "Photography", "Filmmaking"
]

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Scores"

# Headers
headers = ["S. No.", "Event Name", "CW", "Ncrypt", "Exun", "CW Points", "Ncrypt Points", "Exun Points"]
ws.append(headers)

# Fill events
for i, event in enumerate(events, start=1):
    ws.append([i, event, "", "", "", "", "", ""])  # empty placeholders for inputs & formulas

# Add formulas
for row in range(2, len(events) + 2):
    # CW points formula (column F)
    ws[f"F{row}"] = f"=IF(C{row}=1,100,IF(C{row}=2,75,IF(C{row}=3,50,0)))"
    # Ncrypt points formula (column G)
    ws[f"G{row}"] = f"=IF(D{row}=1,100,IF(D{row}=2,75,IF(D{row}=3,50,0)))"
    # Exun points formula (column H)
    ws[f"H{row}"] = f"=IF(E{row}=1,100,IF(E{row}=2,75,IF(E{row}=3,50,0)))"

# Totals row
total_row = len(events) + 2
ws[f"B{total_row}"] = "TOTAL"
ws[f"F{total_row}"] = f"=SUM(F2:F{total_row-1})"
ws[f"G{total_row}"] = f"=SUM(G2:G{total_row-1})"
ws[f"H{total_row}"] = f"=SUM(H2:H{total_row-1})"

# Save workbook
wb.save("Event_Points.xlsx")
print("✅ Event_Points.xlsx has been created!")
